import logging

import openpyxl
from openpyxl.styles import PatternFill, colors

from model.draw_function_data import DrawFunctionData
from model.question_data import QuestionData
from service.resolver_1600 import Resolver1600
from service.resolver_1623 import Resolver1623
from service.resolver_1624 import Resolver1624
from service.resolver_1631 import Resolver1631
from service.resolver_1655 import Resolver1655
from service.resolver_1900 import Resolver1900
from service.resolver_1910 import Resolver1910
from service.resolver_1914 import Resolver1914
from service.resolver_1915 import Resolver1915
from service.resolver_1916 import Resolver1916
from service.resolver_1921 import Resolver1921
from service.resolver_1923 import Resolver1923
from service.resolver_1924 import Resolver1924
from service.resolver_1925 import Resolver1925
from service.resolver_1926 import Resolver1926
from service.resolver_1931 import Resolver1931
from service.resolver_1933 import Resolver1933
from service.resolver_1942 import Resolver1942
from service.resolver_1951 import Resolver1951
from service.resolver_1954 import Resolver1954
from service.resolver_1955 import Resolver1955
from service.resolver_1957 import Resolver1957
from service.resolver_1962 import Resolver1962
from service.resolver_1964 import Resolver1964
from utils.data_utils import DataUtils


class PResolver:
    RESOLVER_POOL = {
        '1600': Resolver1600,
        '1623': Resolver1623,
        '1624': Resolver1624,
        '1631': Resolver1631,
        '1655': Resolver1655,

        '1900': Resolver1900,
        '1910': Resolver1910,
        '1914': Resolver1914,
        '1915': Resolver1915,
        '1916': Resolver1916,
        '1921': Resolver1921,
        '1923': Resolver1923,
        '1924': Resolver1924,
        '1925': Resolver1925,
        '1926': Resolver1926,
        '1931': Resolver1931,
        '1933': Resolver1933,
        '1942': Resolver1942,
        '1951': Resolver1951,
        '1954': Resolver1954,
        '1955': Resolver1955,
        '1957': Resolver1957,
        '1962': Resolver1962,
        '1964': Resolver1964,
    }

    @staticmethod
    def read_excel(excel_file_path: str) -> []:
        print('开始读取Excel中的题目数据...')
        workbook = openpyxl.load_workbook(excel_file_path)
        row_index = 0
        question_pool = {}
        for row in workbook.worksheets[0].rows:
            if row[0].value is not None and row_index > 0:
                if row[0].value.strip() != '':
                    question_key = row[3].value
                    # 获取问题对象，如果不存在那么先创建
                    if question_key not in question_pool:
                        temp_question_data = QuestionData()
                        temp_question_data.row_index = row_index
                        temp_question_data.author = row[0].value
                        temp_question_data.question_type_name = row[1].value
                        temp_question_data.question_level = row[2].value
                        temp_question_data.question_key = row[3].value
                        temp_question_data.question_type_key = str(row[4].value)
                        temp_question_data.dimensionX = int(row[8].value.split(',')[0])
                        temp_question_data.dimensionY = int(row[8].value.split(',')[1])
                        temp_question_data.answer_data = row[9].value
                        temp_question_data.original_data = row[10].value
                        question_pool[question_key] = temp_question_data
                    question_data = question_pool[question_key]
                    # 创建问题绘制函数对象
                    draw_function = DrawFunctionData()
                    draw_function.belong_question = question_data
                    draw_function.function_name = row[5].value
                    draw_function.data = row[6].value
                    draw_function.parameters = row[7].value
                    question_data.draw_function_list.append(draw_function)
            row_index = row_index + 1
        return question_pool.values()

    @staticmethod
    def calculate_answer(question_data: QuestionData, success_callback):
        logging.info('【静态调用】开始计算题目答案：' + question_data.question_key)
        if question_data.question_type_key in PResolver.RESOLVER_POOL:
            resolver = PResolver.RESOLVER_POOL[question_data.question_type_key](question_data)
            resolver.calculate_original_data()
            if question_data.answer_data is None or question_data.answer_data == '':
                try:
                    resolver.calculate_answer(success_callback)
                except:
                    print('解题过程中发生错误，可能是题目录入有误，本题【' + question_data.question_key + '】跳过...')
            else:
                print('题目excel中答案不为空，本题【' + question_data.question_key + '】跳过...')
        else:
            print('暂不支持本题目【' + question_data.question_key + '】的题型: ' + question_data.question_type_key)

    @staticmethod
    def calculate_answer_list(question_data_list: [], success_callback):
        logging.info('start calculate answer list')
        for question in question_data_list:
            PResolver.calculate_answer(question, success_callback)

    @staticmethod
    def write_calculate_result_data_to_excel(excel_file_path: str, question_data_list: []):
        print('开始将计算结果数据写入excel')
        workbook = openpyxl.load_workbook(excel_file_path)
        row_index = 0
        question_mapping = {}
        for question in question_data_list:
            question_mapping[question.question_key] = question
        for row in workbook.worksheets[0].rows:
            if row[0].value is not None and row_index > 0:
                if row[0].value.strip() != '':
                    question_key = row[3].value
                    if question_key in question_mapping:
                        xdata = question_mapping[question_key].get_answer_data_str()
                        xdata2 = []
                        if ',' not in xdata:
                            for x in xdata:
                                xdata2.append(x)
                            workbook.worksheets[0].cell(row_index + 1, 10,
                                                        DataUtils.parse_arr_data_to_comma_str_data(xdata2))
                        else:
                            workbook.worksheets[0].cell(row_index + 1, 10,
                                                        question_mapping[question_key].get_answer_data_str())
                        workbook.worksheets[0].cell(row_index + 1, 11,
                                                    question_mapping[question_key].get_editable_original_data_str())
            row_index = row_index + 1
        workbook.save(excel_file_path)
        print('计算结果Excel写出完毕')

    @staticmethod
    def write_check_rule_data_to_excel(excel_file_path: str, question_data_list: []):
        print('开始将题目校验规则写入excel')
        rule_excel = openpyxl.Workbook()
        style_1 = PatternFill("solid", fgColor=colors.YELLOW)
        style_2 = PatternFill("solid", fgColor=colors.GREEN)
        question_index = 0
        for question in question_data_list:
            sheet = rule_excel.create_sheet(question.question_key)
            sheet.column_dimensions['A'].width = 40.0
            sheet.column_dimensions['B'].width = 30.0
            sheet.column_dimensions['C'].width = 60.0
            sheet.column_dimensions['D'].width = 120.0
            sheet.column_dimensions['E'].width = 30.0
            sheet.cell(1, 1, '题目编号')
            sheet.cell(1, 2, '规则函数名称')
            sheet.cell(1, 3, '规则函数参数')
            sheet.cell(1, 4, '规则函数数据')
            sheet.cell(1, 5, '规则函数排序')
            question_index = question_index + 1
            question_style = style_1 if question_index / 2 == 0 else style_2
            row_index = 2
            for check_rule in question.rules_list:
                sheet.cell(row_index, 1, question.question_key).fill = question_style
                sheet.cell(row_index, 2, check_rule.get_rule_name_str()).fill = question_style
                sheet.cell(row_index, 3, check_rule.get_rule_parameters_str()).fill = question_style
                sheet.cell(row_index, 4, check_rule.get_rule_data_str()).fill = question_style
                sheet.cell(row_index, 5, str(row_index - 2)).fill = question_style
                row_index = row_index + 1
        rule_excel.remove(rule_excel.worksheets[0])
        rule_excel.save(excel_file_path)
        print('题目校验规则excel写出完毕')

    @staticmethod
    def process_question_answer_result_string(answer_data: []) -> str:
        result = str(answer_data)
        result = result.replace(' ', '').replace(',', '').replace('[', '').replace(']', '').replace('\'', '')
        return result
