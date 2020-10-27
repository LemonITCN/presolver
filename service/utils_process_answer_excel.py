import shutil

import openpyxl


class UtilsProcessAnswerExcel:
    @staticmethod
    def process(excel_path: str, answer_dbn_params: str):
        print('开始读取Excel中的题目数据...')
        excel_path_new = excel_path.replace('.xlsx', '_ans.xlsx')
        shutil.copyfile(excel_path, excel_path_new)
        workbook = openpyxl.load_workbook(excel_path_new)
        row_index = 0
        # 题编号 : 题答案
        question_data_pool = {}
        # 题目号 : 答案行索引
        question_location_pool = {}
        question_list = []
        process_question_num = ''
        row_index = 0
        for row in workbook.worksheets[0].rows:
            row_index = row_index + 1
            question_num = row[3].value
            if question_num not in question_num:
                question_data_pool[question_num] = row[9]
            if process_question_num != question_num:
                # 题号改变了
                if process_question_num == '':
                    # 是第一道题
                    process_question_num = question_num
                    print('首题：' + str(question_num))
                else:
                    # 不是第一道题，写入上一道题的数据
                    process_question_num = question_num
                    question_list.append(question_num)
                    row_index = row_index + 1
                    question_location_pool[question_num] = row_index
                    # row_index = row_index + 1
                    print('非首题：' + str(question_num))
        for question_num in question_list:
            new_row_index = question_location_pool[question_num]
            workbook.worksheets[0].insert_rows(new_row_index)

            workbook.worksheets[0].cell(new_row_index, 1, workbook.worksheets[0].cell(new_row_index - 1, 1).value)
            workbook.worksheets[0].cell(new_row_index, 2, workbook.worksheets[0].cell(new_row_index - 1, 2).value)
            workbook.worksheets[0].cell(new_row_index, 3, workbook.worksheets[0].cell(new_row_index - 1, 3).value)
            workbook.worksheets[0].cell(new_row_index, 4, workbook.worksheets[0].cell(new_row_index - 1, 4).value)
            workbook.worksheets[0].cell(new_row_index, 5, workbook.worksheets[0].cell(new_row_index - 1, 5).value)
            workbook.worksheets[0].cell(new_row_index, 6, 'DBN')
            workbook.worksheets[0].cell(new_row_index, 7,
                                        workbook.worksheets[0].cell(new_row_index - 1, 10).value.replace(',', ''))
            workbook.worksheets[0].cell(new_row_index, 8, answer_dbn_params)
            workbook.worksheets[0].cell(new_row_index, 9, workbook.worksheets[0].cell(new_row_index - 1, 9).value)
            print('完善数据：' + question_num)
        workbook.save(excel_path_new)
