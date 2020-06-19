import logging

import openpyxl

from model.question_data import QuestionData
from service.generator_1900 import Generator1900


class PGenerator:
    GENERATOR_POOL = {
        '1900': Generator1900,
    }

    @staticmethod
    def generate(question_seed: QuestionData, question_key: str) -> QuestionData:
        logging.info('【静态调用】开始生成题目：' + question_seed.question_type_key)
        if question_seed.question_type_key in PGenerator.GENERATOR_POOL:
            generator = PGenerator.GENERATOR_POOL[question_seed.question_type_key](question_seed)
            return generator.generate_new_question(question_key)
        else:
            print('暂不支持生成【' + question_seed.question_type_key + '】的题型')
            return QuestionData()

    @staticmethod
    def write_calculate_result_data_to_excel(excel_file_path: str, question_data_list: []):
        print('开始将计算结果数据写入excel')
        workbook = openpyxl.load_workbook(excel_file_path)
        row_index = 1
        for question in question_data_list:
            for draw_function in question.draw_function_list:
                workbook.worksheets[0].cell(row_index + 1, 1, question.author)
                workbook.worksheets[0].cell(row_index + 1, 2, question.question_type_name)
                workbook.worksheets[0].cell(row_index + 1, 3, question.question_level)
                workbook.worksheets[0].cell(row_index + 1, 4, question.question_key)
                workbook.worksheets[0].cell(row_index + 1, 5, question.question_type_key)
                workbook.worksheets[0].cell(row_index + 1, 6, draw_function.function_name)
                workbook.worksheets[0].cell(row_index + 1, 7, draw_function.data)
                workbook.worksheets[0].cell(row_index + 1, 8, draw_function.parameters)
                workbook.worksheets[0].cell(row_index + 1, 9, question.get_dimension_data_str())
                workbook.worksheets[0].cell(row_index + 1, 10, question.get_answer_data_str())
                workbook.worksheets[0].cell(row_index + 1, 11, question.get_original_data_str())
                row_index = row_index + 1
        workbook.save(excel_file_path)
        print('Excel写出完毕')
